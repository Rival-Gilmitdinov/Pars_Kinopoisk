from openpyxl import Workbook, load_workbook

def make_excel_file(filename, data):
    """Функция по соданию таблицы exel
    Arg:
        filename: имя таблицы
        data: данные, добавляемые в таблицу"""
    wb = Workbook()
    ws = wb.active
    ws.title = filename
    fieldnames = ['name', 'year', 'premiere', 'country', 'description', 'rating', 'mpaa', 'adult']
    ws.append(["Название фильма", "Год выпуска", "Премьера", "Страна", "Описание", "Рейтинг", "MPAA", "Ограничение по возрасту"])
    for movie in data:
        values = (movie[k] for k in fieldnames)
        ws.append(values)
    wb.save('films.xlsx')

def write_excel_data(filename: str, data: str, listname: str, fieldnames: list):
    """Функция по добавлению значений в таблицу
    Arg:
        filename: имя таблицы
        data: данные, добавляемые в таблицу
        listname: имя листа
        fieldnames: данные, добавляемые из postgressql"""
    wb = load_workbook(filename)
    ws = wb.create_sheet(listname)
    ws.append(fieldnames)
    for values in data:
        ws.append(values)
    wb.save('films.xlsx')
